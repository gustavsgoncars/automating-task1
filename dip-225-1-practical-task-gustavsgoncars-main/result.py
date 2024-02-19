from openpyxl import Workbook, load_workbook
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
#write your code here
max_row=ws.max_row
for row in range(2,max_row+1):
    hour=ws['B' + str(row)].value
    rate=ws['C' + str(row)].value
    if (type(hour)!=str and type(rate)!=str):
        salary=hour*rate
    if (salary>3000):
       total=total+1
print(total)
wb.close()
